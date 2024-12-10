import os
from openpyxl import Workbook
from openpyxl.styles import Font

def create_index(folder_path, output_excel):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Index"
    bold_font = Font(bold=True)

    row_tracker = []

    def add_to_sheet(level, name, is_folder):
        row = ["" for _ in range(level)]
        row[level - 1] = name
        worksheet.append(row)

        if is_folder:
            cell = worksheet.cell(row=worksheet.max_row, column=level)
            cell.font = bold_font
            row_tracker.append((worksheet.max_row, level))

    for root, dirs, files in os.walk(folder_path):
        depth = root.replace(folder_path, "").count(os.sep) + 1
        folder_name = os.path.basename(root) if depth > 1 else os.path.basename(folder_path)
        add_to_sheet(depth, folder_name, is_folder=True)

        for file_name in files:
            add_to_sheet(depth + 1, file_name, is_folder=False)

    for i in range(len(row_tracker) - 1):
        start_row, level = row_tracker[i]
        
        end_row = worksheet.max_row
        for j in range(i + 1, len(row_tracker)):
            next_row, next_level = row_tracker[j]
            if next_level <= level:
                end_row = next_row - 1
                break

        if end_row > start_row:
            worksheet.row_dimensions.group(start=start_row + 1, end=end_row, outline_level=level, hidden=False)

    workbook.save(output_excel)
    print(f"File saved at {output_excel}")

folder_to_scan = r"Path Here"
output_file = "index.xlsx"

create_index(folder_to_scan, output_file)
